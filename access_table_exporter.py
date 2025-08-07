#!/usr/bin/env python3
"""
Microsoft Access Table Exporter - Professional Version

Exportiert Tabellen aus Microsoft Access-Dateien (.accdb/.mdb) 
in verschiedene Formate mit Fortschrittsanzeige und umfassendem Logging.

Features:
- Standard Input/Output/Log Verzeichnisse
- tqdm Fortschrittsanzeige 
- Umfassendes Logging (Console + Datei)
- GUI-Dialoge mit intelligenten Standardpfaden
- Robuste Fehlerbehandlung
- Export-Zusammenfassung

Autor: GitHub Copilot
Version: 2.0
"""

import sys
import os
import logging
import argparse
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Any, Dict, Tuple
from logging.handlers import RotatingFileHandler

# Third-party imports
try:
    import pyodbc
    import pandas as pd
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from tqdm import tqdm
except ImportError as e:
    print(f"❌ Fehler beim Importieren erforderlicher Bibliotheken: {e}")
    print("Führen Sie aus: pip install pyodbc pandas tqdm")
    sys.exit(1)

# Optional imports
try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️  fpdf2 nicht verfügbar - PDF-Export deaktiviert")

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("⚠️  openpyxl nicht verfügbar - XLSX-Export deaktiviert")

# Konstanten
INPUT_DIR = Path("./input")
OUTPUT_DIR = Path("./export") 
LOG_DIR = Path("./logs")

class ProgressWindow:
    """GUI-Fortschrittsanzeige mit tkinter."""
    
    def __init__(self, total_tables: int):
        self.root = tk.Toplevel()
        self.root.title("Export-Fortschritt")
        self.root.geometry("500x150")
        self.root.resizable(False, False)
        
        # Progress Variables
        self.progress_var = tk.DoubleVar()
        self.total_tables = total_tables
        self.current_table = 0
        
        # UI Elements
        self.setup_ui()
        
    def setup_ui(self):
        """Erstellt die UI-Elemente."""
        # Main Label
        self.main_label = tk.Label(
            self.root, 
            text="Exportiere Access-Tabellen...", 
            font=("Arial", 12, "bold")
        )
        self.main_label.pack(pady=10)
        
        # Current Table Label
        self.table_label = tk.Label(
            self.root, 
            text="Bereit für Export...", 
            font=("Arial", 10)
        )
        self.table_label.pack(pady=5)
        
        # Progress Bar
        self.progress_bar = ttk.Progressbar(
            self.root, 
            variable=self.progress_var, 
            maximum=self.total_tables,
            length=400
        )
        self.progress_bar.pack(pady=10)
        
        # Status Label
        self.status_label = tk.Label(
            self.root, 
            text=f"0 / {self.total_tables} Tabellen", 
            font=("Arial", 9)
        )
        self.status_label.pack(pady=5)
        
    def update(self, table_name: str, success: bool = True):
        """Aktualisiert den Fortschritt."""
        self.current_table += 1
        
        status = "✓" if success else "❌"
        self.table_label.config(text=f"{status} {table_name}")
        self.progress_var.set(self.current_table)
        self.status_label.config(text=f"{self.current_table} / {self.total_tables} Tabellen")
        
        self.root.update()
        
    def close(self):
        """Schließt das Fortschrittsfenster."""
        self.root.destroy()

class AccessTableExporter:
    """Professioneller Microsoft Access Table Exporter."""
    
    def __init__(self, gui_progress: bool = False, chunk_size: Optional[int] = None):
        """
        Initialisiert den Access Table Exporter.
        
        Args:
            gui_progress: Aktiviert GUI-Fortschrittsanzeige
            chunk_size: Chunk-Größe für große Tabellen
        """
        self.gui_progress = gui_progress
        self.chunk_size = chunk_size
        self.logger: Optional[logging.Logger] = None
        self.export_stats: Dict[str, Any] = {
            'start_time': None,
            'end_time': None,
            'total_tables': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'export_details': []
        }
        
        # GUI-Progress Window
        self.progress_window: Optional[ProgressWindow] = None
        
    def setup_dirs(self) -> None:
        """Erstellt erforderliche Verzeichnisse."""
        directories = [INPUT_DIR, OUTPUT_DIR, LOG_DIR]
        
        print("📁 Erstelle Verzeichnisse...")
        for directory in directories:
            try:
                directory.mkdir(parents=True, exist_ok=True)
                print(f"✓ Verzeichnis bereit: {directory}")
            except Exception as e:
                print(f"❌ Fehler beim Erstellen von {directory}: {e}")
                sys.exit(1)
    
    def setup_logger(self) -> None:
        """Konfiguriert das Logging-System."""
        try:
            # Logger erstellen
            self.logger = logging.getLogger("AccessExporter")
            self.logger.setLevel(logging.DEBUG)
            
            # Vorherige Handler entfernen
            for handler in self.logger.handlers[:]:
                self.logger.removeHandler(handler)
            
            # Formatter
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            
            # Console Handler (INFO Level)
            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setLevel(logging.INFO)
            console_handler.setFormatter(formatter)
            
            # File Handler (DEBUG Level, Rotating)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_file = LOG_DIR / f"access_export_{timestamp}.log"
            
            file_handler = RotatingFileHandler(
                log_file, 
                maxBytes=10*1024*1024,  # 10MB
                backupCount=5,
                encoding='utf-8'
            )
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)
            
            # Handler hinzufügen
            self.logger.addHandler(console_handler)
            self.logger.addHandler(file_handler)
            
            self.logger.info("=== ACCESS TABLE EXPORTER GESTARTET ===")
            self.logger.info(f"Log-Datei: {log_file}")
            
        except Exception as e:
            print(f"❌ Logging-Setup fehlgeschlagen: {e}")
            sys.exit(1)
    
    def choose_file(self) -> Optional[Path]:
        """
        Öffnet Dateiauswahl-Dialog für Access-Dateien.
        
        Returns:
            Pfad zur ausgewählten Datei oder None
        """
        try:
            root = tk.Tk()
            root.withdraw()
            
            # Dialog mit INPUT_DIR als Startverzeichnis
            file_path = filedialog.askopenfilename(
                title="Microsoft Access Datei auswählen",
                initialdir=str(INPUT_DIR),
                filetypes=[
                    ("Access-Dateien", "*.accdb *.mdb"),
                    ("Access 2007+", "*.accdb"),
                    ("Access 97-2003", "*.mdb"),
                    ("Alle Dateien", "*.*")
                ]
            )
            
            root.destroy()
            
            if file_path:
                path = Path(file_path)
                self.logger.info(f"Eingabedatei ausgewählt: {path}")
                return path
            else:
                self.logger.warning("Keine Datei ausgewählt")
                return None
                
        except Exception as e:
            self.logger.error(f"Fehler bei Dateiauswahl: {e}")
            return None
    
    def choose_folder(self) -> Optional[Path]:
        """
        Öffnet Ordnerauswahl-Dialog für Ausgabeverzeichnis.
        
        Returns:
            Pfad zum ausgewählten Ordner oder None
        """
        try:
            root = tk.Tk()
            root.withdraw()
            
            # Dialog mit OUTPUT_DIR als Startverzeichnis
            folder_path = filedialog.askdirectory(
                title="Ausgabeordner auswählen",
                initialdir=str(OUTPUT_DIR)
            )
            
            root.destroy()
            
            if folder_path:
                path = Path(folder_path)
                self.logger.info(f"Ausgabeordner ausgewählt: {path}")
                return path
            else:
                self.logger.warning("Kein Ordner ausgewählt")
                return None
                
        except Exception as e:
            self.logger.error(f"Fehler bei Ordnerauswahl: {e}")
            return None
    
    def choose_format(self) -> Optional[str]:
        """
        Lässt Benutzer Exportformat auswählen.
        
        Returns:
            Gewähltes Format oder None
        """
        formats = ['CSV']
        
        if XLSX_AVAILABLE:
            formats.append('XLSX')
        if PDF_AVAILABLE:
            formats.append('PDF')
        
        formats.append('JSON')
        
        print("\n📊 Verfügbare Exportformate:")
        for i, fmt in enumerate(formats, 1):
            print(f"  {i}. {fmt}")
        
        while True:
            try:
                choice = input(f"\nFormat wählen (1-{len(formats)}, oder Name): ").strip()
                
                if choice.upper() in formats:
                    selected = choice.upper()
                    break
                elif choice.isdigit():
                    idx = int(choice) - 1
                    if 0 <= idx < len(formats):
                        selected = formats[idx]
                        break
                
                print("❌ Ungültige Eingabe. Bitte versuchen Sie es erneut.")
                
            except KeyboardInterrupt:
                return None
                
        self.logger.info(f"Exportformat ausgewählt: {selected}")
        return selected
    
    def list_tables(self, db_path: Path) -> List[str]:
        """
        Listet alle Tabellen in der Access-Datenbank auf.
        
        Args:
            db_path: Pfad zur Access-Datei
            
        Returns:
            Liste der Tabellennamen
        """
        tables = []
        connection = None
        
        try:
            self.logger.info(f"Verbinde mit Datenbank: {db_path}")
            
            # ODBC-Treiber probieren
            drivers = [
                "Microsoft Access Driver (*.mdb, *.accdb)",
                "Microsoft Access Driver (*.mdb)",
                "Driver do Microsoft Access (*.mdb)"
            ]
            
            for driver in drivers:
                try:
                    connection_string = f"DRIVER={{{driver}}};DBQ={db_path};"
                    connection = pyodbc.connect(connection_string, timeout=30)
                    self.logger.debug(f"Verbindung erfolgreich mit Treiber: {driver}")
                    break
                except pyodbc.Error:
                    continue
            
            if not connection:
                raise Exception("Kein funktionsfähiger ODBC-Treiber gefunden")
            
            # Tabellen abrufen
            cursor = connection.cursor()
            
            # Methode 1: MSysObjects (falls verfügbar)
            try:
                cursor.execute("SELECT Name FROM MSysObjects WHERE Type=1 AND Flags=0")
                for row in cursor.fetchall():
                    table_name = row[0]
                    if not table_name.startswith(('MSys', 'USys', '~')):
                        tables.append(table_name)
                        
            except pyodbc.Error:
                # Methode 2: cursor.tables() als Fallback
                self.logger.debug("MSysObjects nicht verfügbar, verwende cursor.tables()")
                for table_info in cursor.tables(tableType='TABLE'):
                    table_name = table_info.table_name
                    if not table_name.startswith(('MSys', 'USys', '~')):
                        tables.append(table_name)
            
            tables.sort()
            self.logger.info(f"Gefundene Tabellen ({len(tables)}): {', '.join(tables)}")
            
        except Exception as e:
            self.logger.error(f"Fehler beim Auflisten der Tabellen: {e}")
            self.logger.debug(traceback.format_exc())
            
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
        
        return tables
    
    def select_tables(self, tables: List[str]) -> List[str]:
        """
        Lässt Benutzer Tabellen für Export auswählen.
        
        Args:
            tables: Liste verfügbarer Tabellen
            
        Returns:
            Liste ausgewählter Tabellen
        """
        if not tables:
            print("❌ Keine Tabellen gefunden!")
            return []
        
        if len(tables) == 1:
            print(f"✓ Nur eine Tabelle gefunden: {tables[0]}")
            self.logger.info(f"Automatische Auswahl: {tables[0]}")
            return tables
        
        print(f"\n📋 Gefundene Tabellen ({len(tables)}):")
        for i, table in enumerate(tables, 1):
            print(f"  {i:2d}. {table}")
        
        print("\n🎯 Auswahl-Optionen:")
        print("  - ENTER oder 'alle': Alle Tabellen")
        print("  - Nummern: z.B. '1,3,5' oder '1-3,7'")
        print("  - Namen: z.B. 'Kunden,Bestellungen'")
        
        while True:
            try:
                choice = input("\nTabellen auswählen: ").strip()
                
                if not choice or choice.lower() == 'alle':
                    selected = tables
                    break
                
                selected = []
                
                # Nummern parsen
                if any(c.isdigit() or c in ',-' for c in choice):
                    for part in choice.split(','):
                        part = part.strip()
                        if '-' in part:
                            # Bereich: z.B. "1-5"
                            start, end = map(int, part.split('-'))
                            for i in range(start, end + 1):
                                if 1 <= i <= len(tables):
                                    selected.append(tables[i - 1])
                        elif part.isdigit():
                            # Einzelne Nummer
                            i = int(part)
                            if 1 <= i <= len(tables):
                                selected.append(tables[i - 1])
                else:
                    # Namen parsen
                    names = [name.strip() for name in choice.split(',')]
                    for name in names:
                        if name in tables:
                            selected.append(name)
                
                if selected:
                    break
                else:
                    print("❌ Keine gültigen Tabellen ausgewählt. Bitte versuchen Sie es erneut.")
                    
            except KeyboardInterrupt:
                return []
            except Exception as e:
                print(f"❌ Eingabefehler: {e}")
        
        # Duplikate entfernen und sortieren
        selected = sorted(list(set(selected)))
        
        print(f"✓ Ausgewählt: {len(selected)} Tabelle(n)")
        self.logger.info(f"Ausgewählte Tabellen: {', '.join(selected)}")
        
        return selected
    
    def sanitize_filename(self, filename: str) -> str:
        """
        Bereinigt Dateinamen von ungültigen Zeichen.
        
        Args:
            filename: Ursprünglicher Dateiname
            
        Returns:
            Bereinigter Dateiname
        """
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # Führende/nachfolgende Punkte und Leerzeichen entfernen
        filename = filename.strip('. ')
        
        return filename
    
    def export_table_csv(self, df: pd.DataFrame, output_path: Path) -> None:
        """Exportiert DataFrame als CSV."""
        df.to_csv(
            output_path, 
            index=False, 
            encoding='utf-8-sig', 
            sep=';'
        )
    
    def export_table_xlsx(self, df: pd.DataFrame, output_path: Path) -> None:
        """Exportiert DataFrame als XLSX."""
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl nicht verfügbar")
        
        df.to_excel(
            output_path, 
            index=False, 
            engine='openpyxl'
        )
    
    def export_table_json(self, df: pd.DataFrame, output_path: Path) -> None:
        """Exportiert DataFrame als JSON."""
        df.to_json(
            output_path, 
            orient='records', 
            force_ascii=False, 
            indent=2
        )
    
    def export_table_pdf(self, df: pd.DataFrame, output_path: Path, table_name: str) -> None:
        """Exportiert DataFrame als PDF."""
        if not PDF_AVAILABLE:
            raise ImportError("fpdf2 nicht verfügbar")
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        
        # Titel
        pdf.cell(0, 10, f'Tabelle: {table_name}', 0, 1, 'C')
        pdf.ln(5)
        
        # Spaltenbreite berechnen
        page_width = pdf.w - 2 * pdf.l_margin
        col_count = len(df.columns)
        col_width = page_width / col_count if col_count > 0 else page_width
        
        # Schriftgröße anpassen
        font_size = max(6, min(10, int(60 / col_count))) if col_count > 0 else 10
        
        # Tabellenkopf
        pdf.set_font('Arial', 'B', font_size)
        for column in df.columns:
            pdf.cell(col_width, 8, str(column)[:15], 1, 0, 'C')
        pdf.ln()
        
        # Datenzeilen (begrenzt auf 50 für Performance)
        pdf.set_font('Arial', '', font_size)
        max_rows = min(50, len(df))
        
        for _, row in df.head(max_rows).iterrows():
            for value in row:
                cell_value = str(value) if pd.notna(value) else ''
                pdf.cell(col_width, 6, cell_value[:15], 1, 0, 'C')
            pdf.ln()
        
        # Hinweis bei gekürzten Daten
        if len(df) > max_rows:
            pdf.ln(5)
            pdf.set_font('Arial', 'I', 8)
            pdf.cell(0, 6, f'(Zeigt {max_rows} von {len(df)} Zeilen)', 0, 1, 'C')
        
        pdf.output(str(output_path))
    
    def export_single_table(self, db_path: Path, table_name: str, 
                          output_dir: Path, export_format: str) -> Tuple[bool, str, Dict[str, Any]]:
        """
        Exportiert eine einzelne Tabelle.
        
        Args:
            db_path: Pfad zur Access-Datei
            table_name: Name der Tabelle
            output_dir: Ausgabeverzeichnis
            export_format: Exportformat
            
        Returns:
            Tuple[success, output_file, stats]
        """
        start_time = time.time()
        connection = None
        stats = {
            'table_name': table_name,
            'start_time': datetime.now(),
            'row_count': 0,
            'file_size': 0,
            'duration': 0,
            'success': False,
            'error': None
        }
        
        try:
            self.logger.debug(f"Starte Export von Tabelle: {table_name}")
            
            # Datenbankverbindung
            drivers = [
                "Microsoft Access Driver (*.mdb, *.accdb)",
                "Microsoft Access Driver (*.mdb)",
                "Driver do Microsoft Access (*.mdb)"
            ]
            
            for driver in drivers:
                try:
                    connection_string = f"DRIVER={{{driver}}};DBQ={db_path};"
                    connection = pyodbc.connect(connection_string, timeout=30)
                    break
                except pyodbc.Error:
                    continue
            
            if not connection:
                raise Exception("Keine ODBC-Verbindung möglich")
            
            # Daten lesen
            query = f"SELECT * FROM [{table_name}]"
            
            if self.chunk_size:
                # Chunk-basierter Import für große Tabellen
                chunks = []
                chunk_iterator = pd.read_sql(
                    query, 
                    connection, 
                    chunksize=self.chunk_size
                )
                
                for chunk in tqdm(chunk_iterator, desc=f"Lade {table_name}", leave=False):
                    chunks.append(chunk)
                
                df = pd.concat(chunks, ignore_index=True)
            else:
                # Direkter Import
                df = pd.read_sql(query, connection)
            
            stats['row_count'] = len(df)
            self.logger.debug(f"Tabelle {table_name}: {len(df)} Zeilen geladen")
            
            if df.empty:
                self.logger.warning(f"Tabelle {table_name} ist leer")
                stats['error'] = "Tabelle ist leer"
                return False, "", stats
            
            # Ausgabedatei erstellen
            safe_name = self.sanitize_filename(table_name)
            extension = export_format.lower()
            output_file = output_dir / f"{safe_name}.{extension}"
            
            # Export basierend auf Format
            if export_format == 'CSV':
                self.export_table_csv(df, output_file)
            elif export_format == 'XLSX':
                self.export_table_xlsx(df, output_file)
            elif export_format == 'JSON':
                self.export_table_json(df, output_file)
            elif export_format == 'PDF':
                self.export_table_pdf(df, output_file, table_name)
            else:
                raise ValueError(f"Nicht unterstütztes Format: {export_format}")
            
            # Dateistatistiken
            if output_file.exists():
                stats['file_size'] = output_file.stat().st_size
                stats['success'] = True
                
                duration = time.time() - start_time
                stats['duration'] = duration
                
                self.logger.info(
                    f"✓ {table_name}: {len(df)} Zeilen → "
                    f"{output_file.name} ({stats['file_size']:,} Bytes, {duration:.2f}s)"
                )
                
                return True, str(output_file), stats
            else:
                stats['error'] = "Ausgabedatei wurde nicht erstellt"
                return False, "", stats
                
        except Exception as e:
            error_msg = f"Fehler beim Exportieren von {table_name}: {e}"
            self.logger.error(error_msg)
            self.logger.debug(traceback.format_exc())
            stats['error'] = str(e)
            stats['duration'] = time.time() - start_time
            return False, "", stats
            
        finally:
            if connection:
                try:
                    connection.close()
                except:
                    pass
    
    def export_tables(self, db_path: Path, tables: List[str], 
                     output_dir: Path, export_format: str) -> None:
        """
        Exportiert alle ausgewählten Tabellen mit Fortschrittsanzeige.
        
        Args:
            db_path: Pfad zur Access-Datei
            tables: Liste der zu exportierenden Tabellen
            output_dir: Ausgabeverzeichnis
            export_format: Exportformat
        """
        self.export_stats['start_time'] = datetime.now()
        self.export_stats['total_tables'] = len(tables)
        
        self.logger.info(f"Starte Export von {len(tables)} Tabellen nach {output_dir}")
        self.logger.info(f"Format: {export_format}")
        
        # GUI-Fortschritt initialisieren
        if self.gui_progress:
            self.progress_window = ProgressWindow(len(tables))
        
        # Console-Fortschritt mit tqdm
        progress_bar = tqdm(
            tables, 
            desc="Exportiere Tabellen", 
            unit="Tabelle",
            ncols=80
        )
        
        for table_name in progress_bar:
            progress_bar.set_description(f"Exportiere {table_name}")
            
            success, output_file, stats = self.export_single_table(
                db_path, table_name, output_dir, export_format
            )
            
            # Statistiken aktualisieren
            self.export_stats['export_details'].append(stats)
            
            if success:
                self.export_stats['successful_exports'] += 1
                print(f"  ✓ {table_name} → {Path(output_file).name}")
            else:
                self.export_stats['failed_exports'] += 1
                print(f"  ❌ {table_name}: {stats.get('error', 'Unbekannter Fehler')}")
            
            # GUI-Fortschritt aktualisieren
            if self.progress_window:
                self.progress_window.update(table_name, success)
        
        self.export_stats['end_time'] = datetime.now()
        
        # GUI-Fortschritt schließen
        if self.progress_window:
            time.sleep(1)  # Kurz anzeigen
            self.progress_window.close()
    
    def write_summary(self, output_dir: Path) -> None:
        """
        Schreibt Export-Zusammenfassung in Datei und Konsole.
        
        Args:
            output_dir: Ausgabeverzeichnis für summary.txt
        """
        total_duration = (
            self.export_stats['end_time'] - self.export_stats['start_time']
        ).total_seconds()
        
        # Summary-Text erstellen
        summary_lines = [
            "=" * 60,
            "ACCESS TABLE EXPORT - ZUSAMMENFASSUNG",
            "=" * 60,
            f"Datum/Zeit: {self.export_stats['start_time'].strftime('%Y-%m-%d %H:%M:%S')}",
            f"Gesamtdauer: {total_duration:.2f} Sekunden",
            "",
            f"Tabellen gesamt: {self.export_stats['total_tables']}",
            f"Erfolgreich: {self.export_stats['successful_exports']}",
            f"Fehlgeschlagen: {self.export_stats['failed_exports']}",
            "",
            "DETAIL-ÜBERSICHT:",
            "-" * 60
        ]
        
        total_rows = 0
        total_size = 0
        
        for detail in self.export_stats['export_details']:
            status = "✓" if detail['success'] else "❌"
            error_info = f" ({detail['error']})" if detail['error'] else ""
            
            summary_lines.append(
                f"{status} {detail['table_name']}: "
                f"{detail['row_count']:,} Zeilen, "
                f"{detail['file_size']:,} Bytes, "
                f"{detail['duration']:.2f}s{error_info}"
            )
            
            if detail['success']:
                total_rows += detail['row_count']
                total_size += detail['file_size']
        
        summary_lines.extend([
            "-" * 60,
            f"Gesamt exportierte Zeilen: {total_rows:,}",
            f"Gesamt Dateigröße: {total_size:,} Bytes ({total_size / 1024 / 1024:.2f} MB)",
            f"Durchschnitt pro Tabelle: {total_duration / len(self.export_stats['export_details']):.2f}s",
            "=" * 60
        ])
        
        summary_text = "\n".join(summary_lines)
        
        # In Konsole ausgeben
        print("\n" + summary_text)
        
        # In Datei schreiben
        try:
            summary_file = output_dir / "summary.txt"
            summary_file.write_text(summary_text, encoding='utf-8')
            print(f"\n📄 Zusammenfassung gespeichert: {summary_file}")
            self.logger.info(f"Zusammenfassung geschrieben: {summary_file}")
        except Exception as e:
            self.logger.error(f"Fehler beim Schreiben der Zusammenfassung: {e}")
        
        # Logging-Zusammenfassung
        self.logger.info("=== EXPORT ABGESCHLOSSEN ===")
        self.logger.info(f"Erfolgreiche Exporte: {self.export_stats['successful_exports']}")
        self.logger.info(f"Fehlgeschlagene Exporte: {self.export_stats['failed_exports']}")
        self.logger.info(f"Gesamtdauer: {total_duration:.2f} Sekunden")
    
    def main(self) -> None:
        """Hauptfunktion des Exporters."""
        try:
            print("🚀 Microsoft Access Table Exporter - Professional Version")
            print("=" * 60)
            
            # 1. Verzeichnisse einrichten
            self.setup_dirs()
            
            # 2. Logging konfigurieren
            print("\n📝 Konfiguriere Logging...")
            self.setup_logger()
            
            # 3. Eingabedatei auswählen
            print("\n📂 Wählen Sie die Access-Datei aus...")
            db_path = self.choose_file()
            if not db_path:
                print("❌ Kein Eingabedatei ausgewählt. Programm beendet.")
                return
            
            print(f"✓ Eingabedatei: {db_path}")
            
            # 4. Ausgabeordner auswählen
            print("\n📁 Wählen Sie den Ausgabeordner aus...")
            output_dir = self.choose_folder()
            if not output_dir:
                print("❌ Kein Ausgabeordner ausgewählt. Programm beendet.")
                return
            
            print(f"✓ Ausgabeordner: {output_dir}")
            
            # 5. Exportformat auswählen
            export_format = self.choose_format()
            if not export_format:
                print("❌ Kein Format ausgewählt. Programm beendet.")
                return
            
            print(f"✓ Exportformat: {export_format}")
            
            # 6. Tabellen auflisten
            print(f"\n🔍 Lade Tabellenliste aus {db_path.name}...")
            tables = self.list_tables(db_path)
            
            if not tables:
                print("❌ Keine Tabellen gefunden. Programm beendet.")
                return
            
            # 7. Tabellen auswählen
            selected_tables = self.select_tables(tables)
            if not selected_tables:
                print("❌ Keine Tabellen ausgewählt. Programm beendet.")
                return
            
            # 8. Export durchführen
            print(f"\n⚡ Starte Export von {len(selected_tables)} Tabelle(n)...")
            print(f"Ziel: {output_dir}")
            print(f"Format: {export_format}")
            
            if self.chunk_size:
                print(f"Chunk-Größe: {self.chunk_size:,} Zeilen")
            
            print("-" * 60)
            
            self.export_tables(db_path, selected_tables, output_dir, export_format)
            
            # 9. Zusammenfassung schreiben
            self.write_summary(output_dir)
            
            print(f"\n🎉 Export abgeschlossen!")
            print(f"📁 Ausgabedateien: {output_dir}")
            print(f"📋 Logs: {LOG_DIR}")
            
        except KeyboardInterrupt:
            print("\n❌ Export durch Benutzer abgebrochen.")
            if self.logger:
                self.logger.warning("Export durch Benutzer abgebrochen")
        except Exception as e:
            error_msg = f"Unerwarteter Fehler: {e}"
            print(f"\n❌ {error_msg}")
            if self.logger:
                self.logger.error(error_msg)
                self.logger.debug(traceback.format_exc())
        
        input("\nDrücken Sie ENTER zum Beenden...")

def parse_arguments():
    """Kommandozeilen-Argumente parsen."""
    parser = argparse.ArgumentParser(
        description="Microsoft Access Table Exporter - Professional Version",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python access_table_exporter.py
  python access_table_exporter.py --gui-progress
  python access_table_exporter.py --chunksize 1000 --gui-progress
        """
    )
    
    parser.add_argument(
        '--gui-progress', 
        action='store_true',
        help='Aktiviert GUI-Fortschrittsanzeige (zusätzlich zur Konsole)'
    )
    
    parser.add_argument(
        '--chunksize', 
        type=int,
        help='Chunk-Größe für große Tabellen (Standard: Keine Chunks)'
    )
    
    return parser.parse_args()

if __name__ == "__main__":
    # Kommandozeilen-Argumente parsen
    args = parse_arguments()
    
    # Exporter initialisieren und starten
    exporter = AccessTableExporter(
        gui_progress=args.gui_progress,
        chunk_size=args.chunksize
    )
    
    exporter.main()
